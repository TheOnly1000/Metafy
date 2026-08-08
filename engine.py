#!/usr/bin/env python3
"""Standalone CLI engine — outputs JSON events to stdout.
V2: All steps properly guarded with try/except so no spinner is left spinning."""

import csv
import json
import os
import re
import shutil
import sys
import tempfile
from datetime import datetime

# Bundle path for pip packages (openpyxl, requests) shipped alongside this script
_libs = os.path.join(os.path.dirname(os.path.abspath(__file__)), "python_libs")
if os.path.isdir(_libs):
    sys.path.insert(0, _libs)

import openpyxl
import requests

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.json")
DB_BASE = os.path.join(os.path.expanduser("~"), "Documents", ".DB-nbcu")

_step_counter = 0

def emit(typ, data):
    line = json.dumps({"type": typ, "data": data}, default=str)
    sys.stdout.write(line + "\n")
    sys.stdout.flush()

def emit_step(sid, message, parent=None):
    global _step_counter
    _step_counter += 1
    emit("step", {"id": sid, "seq": _step_counter, "message": message, "parent": parent, "status": "running"})

def emit_step_done(sid):
    emit("step", {"id": sid, "status": "completed"})

def emit_step_error(sid, message):
    emit("step", {"id": sid, "status": "error", "message": message})

def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)

def get_sheet_id(url):
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    return m.group(1) if m else None

def download_xlsx(sid):
    url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(r.content)
    tmp.close()
    return tmp.name

def _read_header_map(ws):
    """Build a column-name → column-index map from row 4 (header row)."""
    HEADER_ALIASES = {
        "asset id": "asset_id",
        "id": "asset_id",
        "title": "title",
        "season": "season",
        "episode": "episode",
        "num segments": "num_segments",
        "segments": "num_segments",
        "series tms id": "series_tms_id",
        "tms id": "series_tms_id",
        "template url": "template_url",
        "url": "template_url",
        "hyperlink": "template_url",
        "date": "date",
        "air date": "date",
    }
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=4, column=col_idx).value
        if raw is None:
            continue
        key = str(raw).strip().lower()
        resolved = HEADER_ALIASES.get(key)
        if resolved and resolved not in col_map:
            # First occurrence wins — "Date" (col 1) is preferred over "Air Date" (col 10)
            col_map[resolved] = col_idx
    return col_map

def parse_main_sheet(path, target_date):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    col_map = _read_header_map(ws)

    def gc(field, row, fallback):
        col = col_map.get(field)
        if col is not None:
            return ws.cell(row=row, column=col).value
        return ws.cell(row=row, column=fallback).value

    matched = []
    for row_idx in range(5, ws.max_row + 1):
        hc = None
        hl_col = col_map.get("template_url") or 22
        hc_cell = ws.cell(row=row_idx, column=hl_col)
        if hc_cell.hyperlink and hc_cell.value:
            hc = hc_cell
        if not hc:
            continue
        rdc = gc("date", row_idx, 1)
        rds = ""
        if rdc:
            if isinstance(rdc, datetime):
                rds = rdc.strftime("%Y-%m-%d")
            elif hasattr(rdc, "strftime"):
                rds = rdc.strftime("%Y-%m-%d")
            else:
                rds = str(rdc)
        if rds != target_date:
            continue
        raw = gc("num_segments", row_idx, 12)
        try:
            ns = int(float(str(raw)))
        except (ValueError, TypeError):
            ns = 0
        matched.append({
            "asset_id": str(gc("asset_id", row_idx, 14) or ""),
            "title": str(gc("title", row_idx, 4) or ""),
            "season": str(gc("season", row_idx, 5) or ""),
            "episode": str(gc("episode", row_idx, 6) or ""),
            "num_segments": ns,
            "series_tms_id": str(gc("series_tms_id", row_idx, 13) or ""),
            "template_url": hc.hyperlink.target,
        })
    wb.close()
    return matched

def parse_template(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    lr = ws.max_row
    d = {
        "description": str(ws.cell(row=lr, column=3).value or ""),
        "long_description": str(ws.cell(row=lr, column=4).value or ""),
        "series_title": str(ws.cell(row=lr, column=5).value or ""),
        "series_tms_id": str(ws.cell(row=lr, column=6).value or ""),
        "tags": str(ws.cell(row=lr, column=7).value or ""),
        "categories": str(ws.cell(row=lr, column=8).value or ""),
        "language": str(ws.cell(row=lr, column=11).value or ""),
    }
    wb.close()
    return d

def clean_en_vivo(text):
    text = re.sub(r'\bEN\s+VIVO\b[\s:]*', '', text)
    text = re.sub(r'[\s:]*\bEN\s+VIVO\b', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'^[\s:;,.]+', '', text).strip()
    return text

def fmt_val(v):
    if v and v != "None":
        v = v.strip()
        try:
            return str(int(float(v)))
        except ValueError:
            return v
    return ""

def generate_rows(aid, title, season, episode, td, nseg, extra, subs):
    total = nseg + extra
    rows = []
    base = {
        "Title": title,
        "description": td["description"],
        "long_description": td["long_description"],
        "series_title": td["series_title"],
        "series_tms_id": td["series_tms_id"],
        "tags": td["tags"],
        "categories": td["categories"],
        "Video Filename": "",
        "Segments": "",
        "Language": td["language"],
        "season": season,
        "episode": episode,
    }
    for i in range(total):
        letter = chr(65 + i)
        mid = f"{aid}{letter}"
        rows.append({**base, "Asset ID": mid})
        for j in range(1, subs + 1):
            rows.append({**base, "Asset ID": f"{mid}{j}"})
    return rows

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--extra-segments", type=int, default=None)
    parser.add_argument("--subsegments", type=int, default=None)
    parser.add_argument("--config-path", default=None)
    args = parser.parse_args()

    global CONFIG_PATH
    if args.config_path:
        CONFIG_PATH = args.config_path

    target_date = args.date
    output_dir = args.output_dir
    config = load_config()

    # Validate date format — prevents path traversal when used in file paths
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", target_date):
        emit("error", "Invalid date format: " + target_date)
        return

    sid = get_sheet_id(config["sheet_url"])
    extra = args.extra_segments if args.extra_segments is not None else config.get("extra_segments", 4)
    subs = args.subsegments if args.subsegments is not None else config.get("subsegments_per_segment", 2)
    cols = config["csv_columns"]

    # ── Schedule download (already guarded) ──
    emit_step("schedule_download", "Downloading schedule sheet from Google Drive")
    try:
        mp = download_xlsx(sid)
    except Exception as e:
        emit_step_error("schedule_download", str(e))
        emit("error", f"Failed: {e}")
        return
    emit_step_done("schedule_download")

    # ── Schedule parse (was unguarded) ──
    emit_step("schedule_parse", "Parsing schedule data for target date")
    try:
        matched = parse_main_sheet(mp, target_date)
    except Exception as e:
        emit_step_error("schedule_parse", str(e))
        emit("error", f"Parse failed: {e}")
        os.unlink(mp)
        return
    emit_step_done("schedule_parse")

    sched_db = os.path.join(DB_BASE, "schedule")
    os.makedirs(sched_db, exist_ok=True)
    with open(os.path.join(sched_db, f"{target_date}.json"), "w", encoding="utf-8") as f:
        json.dump({"date": target_date, "count": len(matched)}, f, default=str)

    os.unlink(mp)

    if not matched:
        emit("status", f"No shows for {target_date}")
        emit("done", None)
        return

    emit("status", f"Found {len(matched)} show(s)")

    dt = datetime.strptime(target_date, "%Y-%m-%d")
    tmpl_db = os.path.join(DB_BASE, "templates")
    month_f = dt.strftime("%B %Y")
    day_f = dt.strftime("%d %B")
    out_dir = os.path.join(output_dir, str(dt.year), month_f, day_f)
    os.makedirs(tmpl_db, exist_ok=True)
    os.makedirs(out_dir, exist_ok=True)

    tcache = {}
    total = len(matched)
    total_rows = 0

    emit("meta", {"total_shows": total, "output_dir": out_dir, "columns": cols})

    for idx, item in enumerate(matched, 1):
        aid = item["asset_id"]
        ns = item["num_segments"]
        if not aid or ns <= 0:
            emit("log", f"[{idx}/{total}] {aid}: skipped")
            emit("progress", {"current": idx, "total": total})
            continue

        asset_step = f"asset_{aid}"
        emit_step(asset_step, f"Processing {aid}: {item['title']} ({ns} segments)")
        emit("asset_start", aid)

        had_error = False
        tsid = get_sheet_id(item["template_url"])

        # ── Template download/parse/clean (all steps properly guarded) ──
        if tsid not in tcache:
            tmpl_step = f"{asset_step}_template"
            emit_step(tmpl_step, "Downloading template sheet", parent=asset_step)
            try:
                tp = download_xlsx(tsid)
                emit_step_done(tmpl_step)
            except Exception as e:
                emit_step_error(tmpl_step, str(e))
                emit("log", f"  ERROR: {e}")
                emit("progress", {"current": idx, "total": total})
                emit_step_done(asset_step)
                continue

            parse_step = f"{asset_step}_parse"
            emit_step(parse_step, "Reading template metadata", parent=asset_step)
            try:
                td = parse_template(tp)
                emit_step_done(parse_step)
            except Exception as e:
                emit_step_error(parse_step, str(e))
                emit("log", f"  ERROR: {e}")
                os.unlink(tp)
                if tsid not in tcache:
                    emit_step_done(asset_step)
                continue

            clean_step = f"{asset_step}_clean"
            emit_step(clean_step, "Cleaning EN VIVO from descriptions", parent=asset_step)
            try:
                for k in ("description", "long_description"):
                    td[k] = clean_en_vivo(td[k])
                emit_step_done(clean_step)
            except Exception as e:
                emit_step_error(clean_step, str(e))
                emit("log", f"  ERROR: {e}")
                os.unlink(tp)
                if tsid not in tcache:
                    emit_step_done(asset_step)
                continue

            tpath = os.path.join(tmpl_db, f"{tsid}.xlsx")
            shutil.copy2(tp, tpath)
            os.unlink(tp)
            tcache[tsid] = td
        else:
            emit("log", f"  Template cached, reusing")

        td = tcache[tsid]

        # ── Generate rows (was unguarded) ──
        gen_step = f"{asset_step}_generate"
        total_segs = ns + extra
        emit_step(gen_step, f"Generating {total_segs} segment rows ({total_segs * (1 + subs)} total)", parent=asset_step)
        try:
            rows = generate_rows(aid, item["title"],
                                 fmt_val(item["season"]), fmt_val(item["episode"]),
                                 td, ns, extra, subs)
            emit_step_done(gen_step)
        except Exception as e:
            emit_step_error(gen_step, str(e))
            emit("log", f"  ERROR generating rows: {e}")
            emit_step_done(asset_step)
            emit("progress", {"current": idx, "total": total})
            continue

        # ── Write CSV (sanitize asset ID to prevent path traversal) ──
        write_step = f"{asset_step}_write"
        safe_aid = re.sub(r"[^\w\-]", "", aid)
        emit_step(write_step, f"Writing {len(rows)} rows to {safe_aid}.csv", parent=asset_step)
        out_path = os.path.join(out_dir, f"{safe_aid}.csv")
        try:
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=cols)
                w.writeheader()
                for r in rows:
                    w.writerow({c: r.get(c, "") for c in cols})
            emit_step_done(write_step)
        except Exception as e:
            emit_step_error(write_step, str(e))
            emit("log", f"  ERROR writing CSV: {e}")
            emit_step_done(asset_step)
            emit("progress", {"current": idx, "total": total})
            continue

        emit("asset_rows", {"asset_id": aid, "rows": rows, "columns": cols, "step_id": write_step})
        emit("log", f"  Wrote {len(rows)} rows")
        total_rows += len(rows)
        emit("progress", {"current": idx, "total": total})
        emit_step_done(asset_step)

    emit("status", "All done!")
    emit("done", {"shows": total, "output_dir": out_dir, "total_rows": total_rows})

if __name__ == "__main__":
    main()
